from django.shortcuts import render
from django.contrib.auth.models import User
from django.views.generic import CreateView,UpdateView,ListView,View,DetailView,DeleteView
from .forms import NewVehicleForm
from .forms import VehicleEditForm
from django.utils import timezone
from django.shortcuts import render,redirect
from datetime import datetime
from .models import Vehicle
from openpyxl import load_workbook
import json
import re
import string
import random

seen = dict()
# Create your views here.
def upload_vehicle_data2(request):
    print(request.tenant.name)
    if request.method == 'POST':
        excel_file = request.FILES["vehicle_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['vehicle']
        excel_data = list()
        header= ['plate_number','engine_number','chassis_number','maker','manufactured_year','vehicle_type'] 

        for row in ws.iter_rows(min_row=2,max_col=6):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)
                
            if not row_content[0]:
                continue
            row_content[1] =  random_string(20)
            row_content[2] =  random_string(20)
            if row_content[2] and row_content[1]: 
                excel_data.append(dict(zip(header,row_content))) 

        for each_record in excel_data:
            try:
                each_record['created_by'] = request.user 
                each_record['vehicle_type'] = each_record['vehicle_type'].upper() 
                Vehicle.objects.create(**each_record);
            except Exception as e:
                print ("error occurred while creating vehicle" + str(e))

        return redirect('vehicles') 
    return render(request,'healthadmin/upload_vehicle_data.html',{ 'tt' : request.tenant })

def random_string(n):
	letters=string.ascii_letters
	random_str = ''.join(random.choice(letters) for i in range(n))
	if random_str in seen.keys():
		random_string(n)
	else:
		seen[random_str]=1
	return random_str
   
class VehicleCreateView(CreateView):
    model=Vehicle
    form_class=NewVehicleForm
    template_name = 'healthadmin/add_vehicle.html' 
    success_url= 'health_vehicles'

    def form_valid(self,form):
        Vehicle = form.save(commit=False)
        Vehicle.created_by = self.request.user
        Vehicle.created_at = timezone.now().strftime('%Y-%m-%d %H:%M:%S.%f') 
        Vehicle.save()
        return redirect('health_vehicles')

class VehicleUpdateView(UpdateView):
    model = Vehicle 
    form_class=VehicleEditForm
    context_object_name= 'vehicle'
    template_name = 'healthadmin/edit_vehicle.html'
    success_url = '/vehicles/'

    def form_valid(self,form):
        Vehicle = form.save(commit=False)
        Vehicle.updated_at = timezone.now()
        Vehicle.updated_by = self.request.user
        Vehicle.save()
        return redirect('vehicles')
   
class VehicleListView(ListView):
    model = Vehicle 
    context_object_name='vehicles'
    template_name = 'healthadmin/vehicles.html'

    def get_queryset(self):
        qs = super(VehicleListView,self).get_queryset().filter(is_active=True).order_by('plate_number')
        return qs

def delete_vehicle(request):
    vehicle = Vehicle.objects.get(pk=request.POST['id'])
    vehicle.is_active = 'f'
    vehicle.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json") 

