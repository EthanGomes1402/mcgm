from django.shortcuts import render
from django.contrib.auth.models import User
from accounts.models import Appuser
from django.http import HttpResponse,JsonResponse
from .models import Bin,Route,Stop_station,Vehicle,Route_schedule,Contractor,Ward_Contractor_Mapping,Vehicle_Garage_Mapping,Installation,Ewd
from common.models import Ward
from django.shortcuts import render,redirect,get_object_or_404
from django.http import Http404
from .forms import NewRouteForm,NewBinForm,NewStopStationForm,NewVehicleForm,NewRouteScheuleForm,NewContractorForm,NewWCMForm,NewVGMForm,NewInstallationForm,NewEwdForm
from .forms import RouteEditForm,BinEditForm,StopStationEditForm,VehicleEditForm,RouteScheduleEditForm,ContractorEditForm,WCMEditForm,VGMEditForm,InstallationEditForm,EwdEditForm
from django.core.serializers import serialize
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin,UserPassesTestMixin
from django.utils.decorators import method_decorator
from django.views.generic import CreateView,UpdateView,ListView,View,DetailView,DeleteView
from django.utils import timezone
from rest_framework.response import Response
from datetime import datetime
from openpyxl import load_workbook
import json
import re
import string
import random
from fastkml import kml
from django.contrib.gis.gdal.datasource import DataSource
from django.contrib.gis.geos import GEOSGeometry,Point,LineString
from django.contrib.gis.gdal import SpatialReference, CoordTransform
from dal import autocomplete
from django.contrib import messages
from shapely import wkt

#floppyform by defaut use srid 3857 to convert it
#explicitley to 4326 we do the trick as follows
gcoord = SpatialReference(3857)
mycoord = SpatialReference(4326)
trans = CoordTransform(gcoord, mycoord)

#class VehicleAutoCompleteView(autocomplete.Select2QuerySetView):
#    def get_queryset(self):
#        qs= Vehicle.objects.all()
#        if self.q:
#            qs.filter(plate_number__istartswith=self.q)
#        return qs
#
#class UserAutoCompleteView(autocomplete.Select2QuerySetView):
#    def get_queryset(self):
#        qs= User.objects.all()
#        if self.q:
#            qs.filter(username__istartswith=self.q)
#        return qs

seen = dict()
@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_vehicle_contractor_mapping(request):
    if request.method == 'POST':
        excel_file = request.FILES["vcm_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['vehicle']
        excel_data = list()
        header= ['vehicle','contractor' ]

        for row in ws.iter_rows(min_row=2,max_col=2):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            if not row_content[1]:
                continue

            try:
                row_content[0] =  Vehicle.objects.get(plate_number=row_content[0])
                row_content[1] =  User.objects.get(id=request.user.id)
                if row_content[0] and row_content[1]:
                    excel_data.append(dict(zip(header,row_content)))
            except Exception as e:
                print(str(e))
            finally:
                pass

        for each_record in excel_data:
            try:
                each_record['created_by'] = request.user
                Vehicle_Contractor_Mapping.objects.create(**each_record);
            except Exception as e:
                print ("error occurred while creating vehicle" + str(e))

        return HttpResponse('---------'+ str(tuple(ws.rows)) + '----------')
    return render(request,'upload_vcm_data.html',{})

def random_string(n):
	letters=string.ascii_letters
	random_str = ''.join(random.choice(letters) for i in range(n))
	if random_str in seen.keys():
		random_string(n)
	else:
		seen[random_str]=1
	return random_str

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_unallocated_bins_from_ward(request):
    ward = Ward.objects.get(pk=request.GET['id'])
    response_data=dict()
    response_data['status'] = 'success'
    response_data['data'] = list( Bin.objects.filter(ward_id=ward).filter(is_active=True).filter(route_id__isnull=True).filter(code__isnull=True).values('id','name') )
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def allocate_bins_to_route(request):
    bins  = list(map(lambda bn : Bin.objects.get(pk=bn) , request.GET['bins'].split(",")))
    route = Route.objects.get(pk=request.GET['route'])
    no_of_bins_rt = route.bins.all().count()

    for bn in bins:
        no_of_bins_rt = no_of_bins_rt + 1
        seq_in_route  = str(no_of_bins_rt).zfill(3)
        bn.code = str(route.code) + '_' + seq_in_route
        bn.save()
        route.bins.add(bn)

    linestring = LineString([ each_bin.bin_location for each_bin in route.bins.all()])
    route.route_fence = linestring
    route.save()

    response_data=dict()
    response_data['status'] = 'success'
    response_data['data'] = 'Bin allocated successfully'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def deallocate_bin_from_route(request):
    bn        = Bin.objects.get(pk=request.GET['id'])
    rt        = bn.route
    bn.code   = None
    bn.save()
    rt.bins.remove(bn)

    if rt.bins.all().count() > 1 :
        linestring = LineString([ each_bin.bin_location for each_bin in rt.bins.all()])
    else:
        linestring = None

    rt.route_fence = linestring

    sequence = 0
    for each_bin in rt.bins.all():
        each_bin.code = None
        sequence = sequence + 1
        bn_code_changed =  rt.code + '_' + str(sequence).zfill(3)
        each_bin.code = bn_code_changed
        each_bin.save()

    rt.save()

    response_data=dict()
    response_data['status'] = 'success'
    response_data['data'] = 'Deallocated bin from route successfully'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_bin_location_from_route(request):
    rt = Route.objects.get(code=request.GET['id'])
    all_bins_from_route = list()
    for each_bin in rt.bins.all():
        each_bin_data = dict()
        each_bin_data['type']     = 'bin'
        each_bin_data['name']     = each_bin.name
        each_bin_data['code']     = each_bin.code
        each_bin_data['location'] = each_bin.bin_location.geojson
        all_bins_from_route.append(each_bin_data)

    rs = rt.route_schedules.all().first()
    if rs and rs.vehicle.vehicle_type in ['MC','LC']:
        mlc         = rs.mlc
        cp          = rs.chkpst
        mlc_centre  = wkt.loads(str(mlc.stop_station_fence.wkt))
        cp_centre   = wkt.loads(str(cp.stop_station_fence.wkt))
        mlc_centre  = Point(float(mlc_centre.centroid.coords[0][0]), float(mlc_centre.centroid.coords[0][1]))
        cp_centre   = Point(float(cp_centre.centroid.coords[0][0]), float(cp_centre.centroid.coords[0][1]))

        route_start = dict()
        route_start['type']     = 'mlc'
        route_start['name']     = mlc.name
        route_start['code']     = mlc.name
        route_start['location'] = mlc_centre.geojson
        all_bins_from_route.insert(0,route_start)

        route_end  = dict()
        route_end['type']     = 'cp'
        route_end['name']       = cp.name
        route_end['code']       = cp.name
        route_end['location']   = cp_centre.geojson
        all_bins_from_route.append(route_end)

    response_data=dict()
    response_data['status'] = 'success'
    response_data['data']   = all_bins_from_route
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_bins_from_route(request):
    rt = Route.objects.get(id=request.GET['id'])

    all_bins_from_route = list()
    for each_bin in rt.bins.all():
        each_bin_data = dict()
        each_bin_data['name'] = each_bin.code
        each_bin_data['id']   = each_bin.id
        all_bins_from_route.append(each_bin_data)

    response_data=dict()
    response_data['status'] = 'success'
    response_data['data'] = all_bins_from_route
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def reorder_bins(request):
    bns = request.GET.getlist('bins[]')
    bins  = list(map(lambda bn : Bin.objects.get(pk=bn) ,bns))
    for sequence,each_bin in enumerate(bins,start=1):
        rt_code,earlier_sequence = each_bin.code.split('_');
        bn_code_changed =  rt_code + '_' + sequence
        each_bin.code = bn_code_changed
        each_bin.save()

#    rt = bn.route
#    bn.route=None
#    bn.save()
#
#    linestring = LineString([ each_bin.bin_location for each_bin in rt.bins.all()])
#    rt.route_fence = linestring
#    rt.save()

    response_data=dict()
    response_data['status'] = 'success'
    response_data['data'] = 'Deallocated bin from route successfully'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

#########################################upload#############################################
@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_bin_data(request):
    if request.method == 'POST':
        excel_file = request.FILES["bin_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['Sheet1']
        excel_data = list()
        header= ['name', 'latitude','longitude','code','tag','bin_location','bin_fence']

        for row in ws.iter_rows(min_row=2,max_col=7):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            if row_content[2] and row_content[1]:
                row_content[2] = re.sub(r'[^\.\d]','',str(row_content[2]))
                row_content[1] = re.sub(r'[^\.\d]','',str(row_content[1]))
                row_content[3] = random_string(6)
                row_content[4] = ''
                try:
                    row_content[5]=Point(float(row_content[2]), float(row_content[1]))
                except:
                    pass
                else:
                    pass
                row_content[6]=''
                excel_data.append(dict(zip(header,row_content)))

        for each_record in excel_data:
            Bin_ward=''
            try:
                Bin_ward = Ward.objects.filter(is_active='t').filter(ward_fence__contains=each_record['bin_location']).get()

                if Bin_ward:
                    each_record['ward'] = Bin_ward
                    Bin.objects.create(**each_record);
                else:
                    messages.error(request, ('Location for bin {0.name} should be within ward '.format(Bin_ward)).upper())

            except Exception as e:
                if Bin_ward:
                    messages.error(request, ('Location for bin {0.name} should be within ward '.format(Bin_ward)).upper())
                else:
                    messages.error(request, ("Location for bin %s should be within ward " %( each_record['name'])).upper())

        return render(request,'swmadmin/upload_bin_data.html',{})
    return render(request,'swmadmin/upload_bin_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_routes_and_bin_data2(request):
    if request.method == 'POST':
        excel_file = request.FILES["bin_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['routes']
        bins = list()
        routes = list()
        routes_from_file= list()
        routes_from_excel= dict()
        all_wards = dict()
        route_failure_geo = 0
        route_failure_ward= 0
        bin_header= ['name','latitude','longitude','code','tag','bin_location','bin_fence','route']
        route_header= ['name','code','route_fence']

        for row in ws.iter_rows(min_row=2,max_col=6):
            row_content = list()
            if row[0].value is not None:
                for cell in row:
                    cell_content = cell.value
                    cell_content = re.sub('-','_',str(cell_content))
                    row_content.append(cell_content)
            else:
                continue

            name,lon,lat = [row_content[2],row_content[4],row_content[3]]
            route_name,route_code,ward = [row_content[0],row_content[0],row_content[5]]

            if route_code not in routes_from_excel:
                routes_from_excel[route_code]=0

            if lat and lon:
                routes_from_excel[route_code]+=1
                lat = re.sub(r'[^\.\d]','',str(lat))
                lon = re.sub(r'[^\.\d]','',str(lon))
                if re.match("[a-z][A-Z]+ ",row_content[1]):
                    rt_for_bin='_'.join(row_content[1].split('_')[:-1])
                    seq_number= row_content[1].split('_').pop().zfill(3)
                    bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                    bin_code = bin_with_seq_number
                    bin_tag = ''
                else:
                    rt_for_bin=row_content[0]
                    seq_number= str(routes_from_excel[route_code]).zfill(3)
                    bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                    bin_code = bin_with_seq_number
                    bin_tag = ''

                try:
                    location =Point(float(lon),float(lat))
                except Exception as e:
                    print(e)
                else:
                    buffer_width = float(5 / 40000000.0 * 360.0)
                    bin_fence = location.buffer(buffer_width)
                    bins.append(dict(zip(bin_header,[name,lat,lon,bin_code,bin_tag,location,bin_fence])))
                    if route_code not in routes_from_file:
                        routes_from_file.append(route_code)
                        routes.append(dict(zip(route_header,[route_name,route_code,''])))

            else:
                continue

        all_routes = list()
        for each_route in routes:
            try:
                rt =  Route.objects.create(**each_route);
                rt.created_by = request.user
                rt.created_at = timezone.now()
                rt.save()
            except Exception as e:
                print ("Error occurred while creating route " + str(e))
            else:
                all_routes.append(rt)

        for each_bin in bins:
            route_for_bin='_'.join(each_bin['code'].split('_')[:-1])
            rt = Route.objects.get(code=route_for_bin)
            try:
                ward_of_bin = Ward.objects.filter(is_active=True).filter(ward_fence__contains=each_bin['bin_location']).get()
                each_bin['route']= rt
                each_bin['ward'] = ward_of_bin
            except Exception as e:
                print (f"Error while getting ward for bin {each_bin['bin_location']}")

            try:
                bn =  Bin.objects.create(**each_bin);
            except Exception as e:
                print ("Error occurred while creating bin " + str(e))
            else:
                pass

        for each_route in all_routes :
            try:
                all_bins_from_route = each_route.bins.all()
                route_path = LineString([bn.bin_location for bn in all_bins_from_route])
                ward_of_route = Ward.objects.filter(is_active=True).filter(ward_fence__contains=route_path).get()
                each_route.route_fence=route_path
                each_route.ward=ward_of_route
                each_route.save()
            except Exception as e:
                geometry_error = 'LineString requires at least 2 points'
                invalid_ward_error ='Ward matching query does not exist'
                ocurred_error = str(e)
                if re.match(geometry_error,ocurred_error):
                    route_failure_geo = route_failure_geo + 1
                    each_route.bins.clear()
                    each_route.delete()

                if re.match(invalid_ward_error,ocurred_error):
                    route_failure_ward = route_failure_ward + 1
                    ward_of_route      = re.sub(r'W$','',each_route.code.split('_')[0])

                    if len(ward_of_route) > 1:
                        ward_of_route      = ward_of_route[:1] + '/' + ward_of_route[1:]
                    rt_ward            = Ward.objects.get(code=ward_of_route)

                    if ward_of_route in all_wards:
                        all_wards[ward_of_route] = all_wards[ward_of_route] + 1
                    else:
                        all_wards[ward_of_route] = 1

                    each_route.ward    = rt_ward
                    each_route.save()
            else:
                pass

        return redirect('routes')
    return render(request,'swmadmin/upload_bin_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_routes_and_bin_data(request):
    if request.method == 'POST':
        excel_file         = request.FILES["bin_excel_file"]
        wb                 = load_workbook(filename = excel_file)
        ws                 = wb['routes']
        bins               = list()
        routes             = list()
        routes_from_file   = list()
        routes_from_excel  = dict()
        all_wards          = dict()
        route_failure_geo  = 0
        route_failure_ward = 0
        bin_header         = ['name','latitude','longitude','code','tag','bin_location','bin_fence','route','ward']
        route_header       = ['name','code','route_fence']
        bins_routes_info   = dict()

        for row in ws.iter_rows(min_row=2,max_col=6):
            row_content = list()
            if row[0].value is not None:
                for cell in row:
                    cell_content = cell.value
                    #just to replace hyphen with underscore all places
                    cell_content = re.sub('-','_',str(cell_content))
                    row_content.append(cell_content)
            else:
                continue

            name_of_bin,lon,lat        = [row_content[2],row_content[4],row_content[3]]
            route_name,route_code,ward = [row_content[0],row_content[0],row_content[5]]
            bin_code,bin_tag           = '',''

            if route_code not in routes_from_excel:
                bins_routes_info[route_code]=dict()
                bins_routes_info[route_code]['status']=1
                bins_routes_info[route_code]['bins'] = list()
                bins_routes_info[route_code]['route_info'] = dict()
                routes_from_excel[route_code]=0

            if lat and lon:
                location =Point(float(lon),float(lat))
                try:
                    #check if bin lies within ward
                    #ward_of_bin = Ward.objects.filter(is_active=True).filter(ward_fence__contains=location).get()
                    ward_of_bin = Ward.objects.get(code=ward)
                    if ward_of_bin.ward_fence.contains(location):
                        routes_from_excel[route_code]+=1
                        lat = re.sub(r'[^\.\d]','',str(lat))
                        lon = re.sub(r'[^\.\d]','',str(lon))
                        if re.match("[a-z][A-Z]+ ",row_content[1]):
                            rt_for_bin='_'.join(row_content[1].split('_')[:-1])
                            seq_number= row_content[1].split('_').pop().zfill(3)
                            bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                            bin_code = bin_with_seq_number
                        else:
                            rt_for_bin=row_content[0]
                            seq_number= str(routes_from_excel[route_code]).zfill(3)
                            bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                            bin_code = bin_with_seq_number

                        buffer_width = float(5 / 40000000.0 * 360.0)
                        bin_fence = location.buffer(buffer_width)
                        bins.append(dict(zip(bin_header,[name_of_bin,lat,lon,bin_code,bin_tag,location,bin_fence,'',ward_of_bin])))
                        bins_routes_info[route_code]['bins'].append(dict(zip(bin_header,[name_of_bin,lat,lon,bin_code,bin_tag,location,bin_fence,'',ward_of_bin])))
                        if route_code not in routes_from_file:
                            routes_from_file.append(route_code)
                            routes.append(dict(zip(route_header,[route_name,route_code,''])))
                            bins_routes_info[route_code]['route_info'] = dict(zip(route_header,[route_name,route_code,'']))
                    else:
                        bins_routes_info[route_code]['status']=0
                        bins_routes_info[route_code]['bins'].clear()
                        continue
                except ObjectDoesNotExist as den:
                    print(den)
            else:
                continue

        all_routes = list()
        for each_route in bins_routes_info:
            if bins_routes_info[each_route]['status'] == 1 and len(bins_routes_info[each_route]['bins']) > 1:
                try:
                    rt =  Route.objects.create(**bins_routes_info[each_route]['route_info']);
                    rt.created_by = request.user
                    rt.created_at = timezone.now()
                    rt.save()
                    for each_bin in bins_routes_info[each_route]['bins']:
                        each_bin['route']= rt
                        bn =  Bin.objects.create(**each_bin)
                    route_path = LineString([each_bin.bin_location for each_bin in rt.bins.all()])
                    ward_of_route = Ward.objects.filter(is_active=True).filter(ward_fence__contains=route_path).get()
                    rt.route_fence=route_path
                    rt.ward=ward_of_route
                    rt.save()
                except Exception as e:
                    print ("Error occurred while creating route " + str(e))
                else:
                    all_routes.append(rt)
            else:
                messages.error(request, f'Failed to create {each_route} either bin outside of ward or route parameters incorrects')

        return render(request,'swmadmin/upload_bin_data.html',{})
    return render(request,'swmadmin/upload_bin_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_routes_and_bin_with_endpoints(request):
    if request.method == 'POST':
        excel_file         = request.FILES["bin_excel_file"]
        wb                 = load_workbook(filename = excel_file)
        ws                 = wb['routes']
        bins               = list()
        routes             = list()
        routes_from_file   = list()
        routes_from_excel  = dict()
        all_wards          = dict()
        route_failure_geo  = 0
        route_failure_ward = 0
        bin_header         = ['name','latitude','longitude','code','tag','bin_location','bin_fence','route','ward']
        route_header       = ['name','code','route_fence']
        bins_routes_info   = dict()

        for row in ws.iter_rows(min_row=2,max_col=6):
            row_content = list()
            if row[0].value is not None:
                for cell in row:
                    cell_content = cell.value
                    #just to replace hyphen with underscore all places
                    cell_content = re.sub('-','_',str(cell_content))
                    row_content.append(cell_content)
            else:
                continue

            name_of_bin,lon,lat        = [row_content[2],row_content[4],row_content[3]]
            route_name,route_code,ward = [row_content[0],row_content[0],row_content[5]]
            bin_code,bin_tag           = [row_content[1],'']

            if route_code not in routes_from_excel:
                bins_routes_info[route_code]               = dict()
                bins_routes_info[route_code]['status']     = 1
                bins_routes_info[route_code]['start']      = 1
                bins_routes_info[route_code]['end']        = 1
                bins_routes_info[route_code]['bins']       = list()
                bins_routes_info[route_code]['route_info'] = dict()
                routes_from_excel[route_code]              = 0

            if lat and lon:
                location =Point(float(lon),float(lat))
                try:
                    #check if bin lies within ward
                    #ward_of_bin = Ward.objects.filter(is_active=True).filter(ward_fence__contains=location).get()
                    ward_of_bin = Ward.objects.get(code=ward)
                    if ward_of_bin.ward_fence.contains(location):
                        routes_from_excel[route_code]+=1
                        lat = re.sub(r'[^\.\d]','',str(lat))
                        lon = re.sub(r'[^\.\d]','',str(lon))
                        if re.match("[a-z][A-Z]+ ",row_content[1]):
                            rt_for_bin='_'.join(row_content[1].split('_')[:-1])
                            seq_number= row_content[1].split('_').pop().zfill(3)
                            bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                            bin_code = bin_with_seq_number
                        else:
                            rt_for_bin=row_content[0]
                            seq_number= str(routes_from_excel[route_code]).zfill(3)
                            bin_with_seq_number= str(rt_for_bin) + '_' + str(seq_number)
                            bin_code = bin_with_seq_number

                        buffer_width = float(5 / 40000000.0 * 360.0)
                        bin_fence = location.buffer(buffer_width)
                        bins.append(dict(zip(bin_header,[name_of_bin,lat,lon,bin_code,bin_tag,location,bin_fence,'',ward_of_bin])))
                        bins_routes_info[route_code]['bins'].append(dict(zip(bin_header,[name_of_bin,lat,lon,bin_code,bin_tag,location,bin_fence,'',ward_of_bin])))
                        if route_code not in routes_from_file:
                            routes_from_file.append(route_code)
                            routes.append(dict(zip(route_header,[route_name,route_code,''])))
                            bins_routes_info[route_code]['route_info'] = dict(zip(route_header,[route_name,route_code,'']))
                    else:
                        bins_routes_info[route_code]['status']=0
                        bins_routes_info[route_code]['bins'].clear()
                        continue
                except ObjectDoesNotExist as den:
                    print(den)
            else:
                if bin_code.endswith('SS'):
                    ward_of_mlc = Ward.objects.get(code=ward)
                    start_point = Stop_station.objects.filter(is_mlc=True).filter(name=name_of_bin)
                    mlc_centre  = wkt.loads(str(start_point.stop_station_fence.wkt))
                    mlc_centre  = Point(float(mlc_centre.centroid.coords[0][0]), float(mlc_centre.centroid.coords[0][1]))
                    if ward_of_mlc.ward_fence.contains(start_point.stop_station_fence):
                        bins_routes_info[route_code]['start'] = mlc_centre
                    else:
                        bins_routes_info[route_code]['status']=0

                if bin_code.endswith('EE'):
                    ward_of_cp = Ward.objects.get(code=ward)
                    end_point = Stop_station.objects.filter(is_chkpst=True).filter(name=name_of_bin)
                    cp_centre  = wkt.loads(str(end_point.stop_station_fence.wkt))
                    cp_centre  = Point(float(cp_centre.centroid.coords[0][0]), float(cp_centre.centroid.coords[0][1]))
                    if ward_of_cp.ward_fence.contains(end_point.stop_station_fence):
                        bins_routes_info[route_code]['end'] = cp_centre
                    else:
                        bins_routes_info[route_code]['status']=0

        all_routes = list()
        for each_route in bins_routes_info:
            if bins_routes_info[each_route]['status'] == 1 and len(bins_routes_info[each_route]['bins']) > 1:
                try:
                    rt =  Route.objects.create(**bins_routes_info[each_route]['route_info']);
                    rt.created_by = request.user
                    rt.created_at = timezone.now()
                    rt.save()
                    for each_bin in bins_routes_info[each_route]['bins']:
                        each_bin['route']= rt
                        bn =  Bin.objects.create(**each_bin)
                    route_path = LineString([ each_bin.bin_location for each_bin in rt.bins.all()])
                    #route_path = LineString([ bins_routes_info[route_code]['start'],each_bin.bin_location for each_bin in rt.bins.all(),bins_routes_info[route_code]['end']])

                    ward_of_route = Ward.objects.filter(is_active=True).filter(ward_fence__contains=route_path).get()
                    rt.route_fence=route_path
                    rt.ward=ward_of_route
                    rt.save()
                except Exception as e:
                    print ("Error occurred while creating route " + str(e))
                else:
                    all_routes.append(rt)
            else:
                messages.error(request, f'Failed to create {each_route} either bin outside of ward or route parameters incorrects')

        return render(request,'swmadmin/upload_bin_data.html',{})
    return render(request,'swmadmin/upload_bin_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_stop_station_data(request):
    if request.method == 'POST':
        mykmlfile = request.FILES["excel_file"]
        with open('/home/mcgm/Development/mcgm/mcgm/fulldata/stop_stations.kml','wb+') as destination:
            for chunk in mykmlfile.chunks():
                destination.write(chunk)

        ds = DataSource('/home/mcgm/Development/mcgm/mcgm/fulldata/stop_stations.kml')
        for layer in ds:
            for feat in layer:
                st_name = feat.get('name')
                geom = feat.geom
                if layer.name == 'MLC':
                    try:
                        st_ward = Ward.objects.get(code=feat.get('ward'))
                    except Exception as e:
                        print ("error occurred while adding installation" + str(e))

                    st = Stop_station.objects.create(
                        name   = st_name,
                        is_mlc = True,
                        ward = st_ward,
                        stop_station_fence = geom.geos,
                        created_by = request.user,
                        created_at = timezone.now()
                        )

                elif layer.name == 'CHKPST':
                    try:
                        st_ward = Ward.objects.get(code=feat.get('ward'))
                    except Exception as e:
                        print ("error occurred while adding installation" + str(e))

                    st = Stop_station.objects.create(
                        name   = st_name,
                        is_chkpst = True,
                        ward = st_ward,
                        stop_station_fence = geom.geos,
                        created_by = request.user,
                        created_at = timezone.now()
                        )
                elif layer.name == 'TNSSTN':
                    st = Stop_station.objects.create(
                        name   = st_name,
                        is_tnsstn = True,
                        stop_station_fence = geom.geos,
                        created_by = request.user,
                        created_at = timezone.now()
                        )
                elif layer.name == 'DMPGND':
                    st = Stop_station.objects.create(
                        name   = st_name,
                        is_dmpgnd = True,
                        stop_station_fence = geom.geos,
                        created_by = request.user,
                        created_at = timezone.now()
                        )
                elif layer.name == 'GARAGE':
                    st = Stop_station.objects.create(
                        name        = st_name,
                        is_garage   = True,
                        stop_station_fence = geom.geos,
                        created_by = request.user,
                        created_at = timezone.now()
                        )
                else:
                    pass

        return redirect('stop_stations')
    return render(request,'swmadmin/upload_stop_stations.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_vehicle_data(request):
    if request.method == 'POST':
        excel_file = request.FILES["vehicle_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['vehicle']
        excel_data = list()
        header= ['plate_number','engine_number','chassis_number','maker','manufactured_year','vehicle_type','contractor','ward']

        for row in ws.iter_rows(min_row=2,max_col=8):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            row_content[1] =  random_string(20)
            row_content[2] =  random_string(20)
            row_content[0] =  row_content[0].strip()
            row_content[3] =  row_content[3].strip()
            row_content[5] =  row_content[5].strip()
            row_content[6] =  row_content[6].strip()
            row_content[7] =  row_content[7].strip()

            if row_content[2] and row_content[1] and row_content[6]:
                user_cntr  = User.objects.get(username=row_content[6])
                row_content[6] = user_cntr.appuser.bmc_contractor

                if row_content[7]:
                    row_content[7] = Ward.objects.get(code=row_content[7])

                excel_data.append(dict(zip(header,row_content)))

        for each_record in excel_data:
            try:
                each_record['created_by'] = request.user
                each_record['vehicle_type'] = each_record['vehicle_type'].upper()
                Vehicle.objects.create(**each_record);
            except:
                print ("error occurred while creating vehicle")

        return redirect('vehicles')
    return render(request,'swmadmin/upload_vehicle_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_installation_data(request):
    if request.method == 'POST':
        excel_file = request.FILES["installation_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['installation']
        installation_data = list()
        header= ['vehicle','imei','sim','wnld_tag']

        for row in ws.iter_rows(min_row=1,max_col=4):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            row_content[0]=row_content[0].strip()
            row_content[1]=str(row_content[1]).strip()
            row_content[2]=str(row_content[2]).strip()
            row_content[3]=row_content[3].strip()

            if row_content[0] and row_content[1] and row_content[2]:
                try:
                    vehicle = Vehicle.objects.get(plate_number=row_content[0])
                except Vehicle.DoesNotExist:
                    vehicle = None

                if not vehicle:
                    print(f"vehice {row_content[0]} does not exists")
                    continue
                row_content[0] = vehicle
                installation_data.append(dict(zip(header,row_content)))

        for each_installation in installation_data:
            try:
                each_installation['created_by'] = request.user
                Installation.objects.create(**each_installation);
            except Exception as e:
                print ("error occurred while adding installation" + str(e))

        #print(installation_data)
        return redirect('installations')
    return render(request,'swmadmin/upload_installation_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_route_schedule_data(request):
    if request.method == 'POST':
        excel_file = request.FILES["route_schedule_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['route_schedule']
        route_schedule_data = list()
        header= ['name','route','vehicle','shift','mlc','chkpst']


        for row in ws.iter_rows(min_row=2,max_col=6):
            ward_of_rs,vehicle,cp,mlc,rt = None,None,None,None,None
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            try:
                ward_of_route = (row_content[0].split("_"))[0][:-1]
                if len(ward_of_route) == 1:
                    ward_of_rs = Ward.objects.get(code=ward_of_route)
                else:
                    ward_of_route = ward_of_route[:1] + '/' + ward_of_route[1:]
                    ward_of_rs = Ward.objects.get(code=ward_of_route)
            except Ward.DoesNotExist:
                continue

            if row_content[1] and row_content[2] and row_content[3] and  row_content[4] and row_content[5]:
                try:
                    vehicle = Vehicle.objects.get(plate_number=row_content[2])
                    row_content[2] = vehicle
                except Vehicle.DoesNotExist:
                    continue

                try:
                    rt = Route.objects.get(code=row_content[1])
                    row_content[1] = rt
                except Route.DoesNotExist:
                    continue

                try:
                    mlc = Stop_station.objects.get(name=row_content[4])
                    row_content[4] = mlc
                except Stop_station.DoesNotExist:
                    continue

                try:
                    cp = Stop_station.objects.get(name=row_content[5])
                    row_content[5] = cp
                except Stop_station.DoesNotExist:
                    continue

                if len(set([cp.ward,mlc.ward,rt.ward,vehicle.ward])) > 1:
                    messages.error(request,f'for {row_content[0]} all route,vehicle,mlc,checkpost should be within a ward')
                    continue
                else:
                    if ward_of_rs.ward_fence.contains(rt.route_fence) and ward_of_rs.ward_fence.contains(mlc.stop_station_fence) and ward_of_rs.ward_fence.contains(cp.stop_station_fence):
                        pass
                    else:
                        messages.error(request, 'all route,vehicle,mlc,checkpost should be within a ward')
                        continue

                route_schedule_data.append(dict(zip(header,row_content)))

        for each_route_schedule in route_schedule_data:
            try:
                each_route_schedule['created_by'] = request.user
                each_route_schedule['created_at']= timezone.now()
                Route_schedule.objects.create(**each_route_schedule);
            except Exception as e:
                print ("error occurred while adding route schedule" + str(e))
        messages.success(request, 'all routes schedules added successfuly')
        return render(request,'swmadmin/upload_route_schedule_data.html',{})
        #return redirect('route_schedules')
    return render(request,'swmadmin/upload_route_schedule_data.html',{})

@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_geos(request):
    if request.method == 'POST':
        mykmlfile = request.FILES["excel_file"]
        with open('/home/mcgm/Development/mcgm/mcgm/fulldata/geos.kml','wb+') as destination:
            for chunk in mykmlfile.chunks():
                destination.write(chunk)

        ds = DataSource('/home/mcgm/Development/mcgm/mcgm/fulldata/geos.kml')
        print ('--------'+ str(ds.layer_count) + '-------')
        print ('--------'+ str(ds[0].num_feat) + '--------')
        print ('--------'+ str(ds[0].geom_type.name) + '--------')
        #print ('--------'+ str(ds[0].srs) + '--------')
        for layer in ds:
            for feat in layer:
                geom = feat.geom
                print(feat.get('name'))
                print(feat.fields)
                print ("<<<<<<<<<<<<<<<<<<<")
                print(feat.geom_type)
                pass
                #ward.objects.create(
                #    name='abc',
                #    code='hhbhb',
                #    ward_fence=geom.geos  # <- and here
                #)
#            print ("layer =>" + str(layer.name))
#            geoms = layer.get_geoms(geos=True)
#            print ("No of fetaures" + str(len(geoms)))
#            for each_geom in geoms:
#                if each_geom.geom_typeid == 6:
#                    print ('-------------------------')
#                    print ('-------------------------')
#                    print (each_geom.dims)
#                    print (each_geom.geom_type)
#                    print (each_geom.geom_typeid)
#                    print (each_geom.json)
#                    #print (each_geom)
            pass
    return render(request,'upload_student.html',{})

#ward code and contractor username excel sheet
@login_required
@user_passes_test(lambda user: user.is_superuser)
def upload_ward_contractor_mapping(request):
    if request.method == 'POST':
        excel_file = request.FILES["wcm_excel_file"]
        wb = load_workbook(filename = excel_file)
        ws = wb['ward']
        excel_data = list()
        header= ['ward','contractor' ]

        for row in ws.iter_rows(min_row=2,max_col=2):
            row_content = list()
            for cell in row:
                row_content.append(cell.value)

            if not row_content[0]:
                continue

            if not row_content[1]:
                continue

            try:
                row_content[0] =  Ward.objects.get(code=row_content[0])
                row_content[1] =  Contractor.objects.get(name=row_content[1])
                if row_content[0] and row_content[1]:
                    excel_data.append(dict(zip(header,row_content)))
            except Exception as e:
                print(str(e))
            finally:
                pass

        for each_record in excel_data:
            try:
                each_record['created_by'] = request.user
                each_record['created_at']= timezone.now()
                Ward_Contractor_Mapping.objects.create(**each_record);
            except Exception as e:
                print ("error occurred while creating ward" + str(e))

        return redirect('wcms')
    return render(request,'swmadmin/upload_wcm_data.html',{})

def upload_ewd_data(request):
    if request.method == 'POST':
        mykmlfile = request.FILES["excel_file"]
        with open('/home/mcgm/Development/mcgm/mcgm/fulldata/ewds.kml','wb+') as destination:
            for chunk in mykmlfile.chunks():
                destination.write(chunk)

        ds = DataSource('/home/mcgm/Development/mcgm/mcgm/fulldata/ewds.kml')
        for layer in ds:
            for feat in layer:
                #geom = feat.geom.clone()
                geom = feat.geom
                geom.coord_dim = 2
                ewd  = Ewd.objects.create(
                     name        = feat.get('name'),
                     code        = feat.get('name'),
                     ewd_fence   = geom.geos,
                     created_by  = request.user,
                     created_at  = timezone.now()
                )
        return redirect('ewds')
    return render(request,'swmadmin/upload_ewds.html',{})

#########################################Ewd############################################Done
class EwdCreateView(CreateView):
    model=Ewd
    form_class=NewEwdForm
    template_name = 'swmadmin/add_ewd.html'
    success_url = '/ewds/'

    def form_valid(self,form):
        ewd = form.save(commit=False)
        ewd.created_by = self.request.user
        ewd.created_at = timezone.now()
        ewd.save()
        return redirect('ewds')

class EwdUpdateView(UpdateView):
    model = Ewd
    form_class=EwdEditForm
    context_object_name= 'ewd'
    template_name = 'swmadmin/edit_ewd.html'
    success_url = '/ewds/'

    def form_valid(self,form):
        ewd = form.save(commit=False)
        ewd.updated_at = timezone.now()
        ewd.updated_by = self.request.user
        ewd.save()
        return redirect('ewds')

class EwdListView(ListView):
    model = Ewd
    context_object_name='ewds'
    template_name = 'swmadmin/ewds.html'

    def get_queryset(self):
        qs = super(EwdListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

########################################StopStation##########################################Done
class StopStationCreateView(UserPassesTestMixin,CreateView):
    model=Stop_station
    form_class=NewStopStationForm
    template_name = 'swmadmin/add_stop_station.html'
    success_url= '/stop_stations'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Stop_station = form.save(commit=False)
        Stop_station.created_by = self.request.user
        Stop_station.created_at = timezone.now()
        Stop_station.save()
        return redirect('stop_stations')

class StopStationUpdateView(UserPassesTestMixin,UpdateView):
    model = Stop_station
    form_class = StopStationEditForm
    context_object_name= 'stop_station'
    template_name = 'swmadmin/edit_stop_station.html'
    success_url = '/stop_stations/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Stop_station = form.save(commit=False)
        Stop_station.updated_at = timezone.now()
        Stop_station.updated_by = self.request.user
        Stop_station.save()
        return redirect('stop_stations')

class StopStationListView(UserPassesTestMixin,ListView):
    model = Stop_station
    context_object_name='stop_stations'
    template_name = 'swmadmin/stop_stations.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(StopStationListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

####################################Bin#####################################################Done

class BinCreateView(UserPassesTestMixin,CreateView):
    model=Bin
    form_class=NewBinForm
    template_name = 'swmadmin/add_bin.html'
    success_url= '/bins'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Bin = form.save(commit=False)
        Bin.bin_location.transform(trans)
        buffer_width = float(5 / 40000000.0 * 360.0)
        Bin.bin_fence = Bin.bin_location.buffer(buffer_width)
        Bin.latitude = Bin.bin_location.coords[1]
        Bin.longitude = Bin.bin_location.coords[0]
        Bin.ward = Ward.objects.filter(is_active=True).filter(ward_fence__contains=Bin.bin_location).get()
        Bin.created_by = self.request.user
        Bin.created_at = timezone.now()
        Bin.save()
        return redirect('bins')

class BinUpdateView(UserPassesTestMixin,UpdateView):
    model = Bin
    form_class=BinEditForm
    context_object_name= 'bin'
    template_name = 'swmadmin/edit_bin.html'
    success_url = '/bins/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Bin = form.save(commit=False)
        Bin.bin_location.transform(trans)
        buffer_width = float(5 / 40000000.0 * 360.0)
        Bin.bin_fence = Bin.bin_location.buffer(buffer_width)
        Bin.latitude = Bin.bin_location.coords[1]
        Bin.longitude = Bin.bin_location.coords[0]
        Bin.ward = Ward.objects.filter(is_active=True).filter(ward_fence__contains=Bin.bin_location).get()
        Bin.updated_at = timezone.now()
        Bin.updated_by = self.request.user
        Bin.save()
        return redirect('bins')

class BinDeleteView(UserPassesTestMixin,DeleteView):
    model = Bin
    success_url = '/bins/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Bin = form.save(commit=False)
        Bin.bin_location.transform(trans)
        buffer_width = float(5 / 40000000.0 * 360.0)
        Bin.bin_fence = Bin.bin_location.buffer(buffer_width)
        Bin.latitude = Bin.bin_location.coords[1]
        Bin.longitude = Bin.bin_location.coords[0]
        Bin.ward = Ward.objects.filter(is_active=True).filter(ward_fence__contains=Bin.bin_location).get()
        Bin.updated_at = timezone.now()
        Bin.updated_by = self.request.user
        Bin.save()
        return redirect('bins')

class BinListView(UserPassesTestMixin,ListView):
    model = Bin
    context_object_name='bins'
    template_name = 'swmadmin/bins.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(BinListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

#################################RouteSchedule

class RouteScheduleCreateView(UserPassesTestMixin,CreateView):
    model=Route_schedule
    form_class=NewRouteScheuleForm
    template_name = 'swmadmin/add_route_schedule.html'
    success_url= '/route_schedules/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        route_schedule = form.save(commit=False)
        route_ward   = form.cleaned_data['route'].ward
        vehicle_ward = form.cleaned_data['vehicle'].ward
        mlc_ward     = form.cleaned_data['mlc'].ward
        chkpst_ward  = form.cleaned_data['chkpst'].ward

        if len(set([route_ward, vehicle_ward,mlc_ward, chkpst_ward])) > 1:
            form.add_error(None,'Route,Vehicle,MLC and Checkpost must have same ward')
            return super(RouteScheduleCreateView, self).form_invalid(form)

        route_schedule.created_at = timezone.now()
        route_schedule.created_by = self.request.user
        route_schedule.save()
        return redirect('route_schedules')

class RouteScheduleUpdateView(UserPassesTestMixin,UpdateView):
    model = Route_schedule
    form_class = RouteScheduleEditForm
    context_object_name= 'route_schedule'
    template_name = 'swmadmin/edit_route_schedule.html'
    success_url = '/route_schedules/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Route_schedule = form.save(commit=False)
        route_ward   = form.cleaned_data['route'].ward
        vehicle_ward = form.cleaned_data['vehicle'].ward
        mlc_ward     = form.cleaned_data['mlc'].ward
        chkpst_ward  = form.cleaned_data['chkpst'].ward

        if len(set([route_ward, vehicle_ward,mlc_ward, chkpst_ward])) > 1:
            form.add_error(None,'Route,Vehicle,MLC and Checkpost must have same ward')
            return super(RouteScheduleUpdateView, self).form_invalid(form)
        Route_schedule.updated_at = timezone.now()
        Route_schedule.updated_by = self.request.user
        Route_schedule.save()
        return redirect('route_schedules')

class RouteScheduleListView(UserPassesTestMixin,ListView):
    model = Route_schedule
    context_object_name='route_schedules'
    template_name = 'swmadmin/route_schedules.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(RouteScheduleListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

###################################Route#############################################################

class RouteCreateView(UserPassesTestMixin,CreateView):
    model=Route
    form_class=NewRouteForm
    template_name = 'swmadmin/add_route.html'
    success_url= '/routes'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Route = form.save(commit=False)
        Route.created_by = self.request.user
        Route.created_at = timezone.now()
        Route.route_fence.transform(trans)

        if Route.route_type == 'N':
            if Ward.objects.filter(ward_fence__contains=Route.route_fence).count() == 1:
                route_ward = Ward.objects.filter(ward_fence__contains=Route.route_fence).get()
                Route.ward = route_ward
                Route.save()
            else:
                messages.error(self.request, ('Route should be within ward').upper())
                return self.render_to_response(self.get_context_data(form=form))

            if Route.route_fence:
                for sequence,point in enumerate(Route.route_fence.coords,start=1):
                    bin_data= dict()
                    bin_data['name'] = str(Route.code) + '_' + str(f'{sequence:04}')
                    bin_data['code'] = str(Route.code) + '_' + str(f'{sequence:03}')
                    bin_data['latitude']  = point[0]
                    bin_data['longitude']  = point[1]
                    bin_data['bin_location'] = Point(point)
                    buffer_width = float(5 / 40000000.0 * 360.0)
                    bin_data['bin_fence'] = bin_data['bin_location'].buffer(buffer_width)
                    bin_data['route'] = Route
                    bin_data['ward']  = Route.ward
                    BIN =  Bin.objects.create(**bin_data);
        else:
            Route.save()
            if Route.route_fence:
                for sequence,point in enumerate(Route.route_fence.coords,start=1):
                    bin_data= dict()
                    bin_data['name'] = str(Route.code) + '_' + str(f'{sequence:04}')
                    bin_data['code'] = str(Route.code) + '_' + str(f'{sequence:03}')
                    bin_data['latitude']  = point[0]
                    bin_data['longitude']  = point[1]
                    bin_data['bin_location'] = Point(point)
                    buffer_width = float(5 / 40000000.0 * 360.0)
                    bin_data['bin_fence'] = bin_data['bin_location'].buffer(buffer_width)
                    bin_data['route'] = Route
                    bin_data['ward']  = None
                    BIN =  Bin.objects.create(**bin_data);

        return redirect('routes')

class RouteUpdateView(UserPassesTestMixin,UpdateView):
    model = Route
    form_class=RouteEditForm
    context_object_name= 'route'
    template_name = 'swmadmin/edit_route.html'
    success_url = '/routes/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Route = form.save(commit=False)
        Route.updated_at = timezone.now()
        Route.updated_by = self.request.user
        Route.route_fence.transform(trans)

        if Ward.objects.filter(ward_fence__contains=Route.route_fence).count() == 1:
            route_ward = Ward.objects.filter(ward_fence__contains=Route.route_fence).get()
            Route.ward = route_ward
            Route.save()
        else:
            messages.error(self.request, ('Route should be within ward').upper())
            return self.render_to_response(self.get_context_data(form=form))

        if Route.route_fence:
            pass
#            for sequence,point in enumerate(Route.route_fence.coords,start=1):
#                bin_data= dict()
#                bin_data['name'] = str(Route.code) + '_' + str(sequence)
#                bin_data['code'] = str(Route.code) + '_' + str(sequence)
#                bin_data['latitude']  = point[0]
#                bin_data['longitude']  = point[1]
#                bin_data['bin_location'] = Point(point)
#                buffer_width = float(5 / 40000000.0 * 360.0)
#                bin_data['bin_fence'] = bin_data['bin_location'].buffer(buffer_width)
#                bin_data['route'] = Route
#                bin_data['ward']  = Route.ward
#                BIN =  Bin.objects.create(**bin_data);

        return redirect('routes')

class RouteListView(UserPassesTestMixin,ListView):
    model = Route
    context_object_name='routes'
    template_name = 'swmadmin/routes.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(RouteListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

    def get_context_data(self, *args, **kwargs):
        # Call the base implementation first to get a context
        context = super(RouteListView, self).get_context_data(*args, **kwargs)
        # add whatever to your context:
        context['whatever'] = messages
        return context


####################################Vehicle#######################################################

class VehicleCreateView(UserPassesTestMixin,CreateView):
    model=Vehicle
    form_class=NewVehicleForm
    template_name = 'swmadmin/add_vehicle.html'
    success_url= '/vehicles'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Vehicle = form.save(commit=False)
        Vehicle.created_by = self.request.user
        Vehicle.created_at = timezone.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        Vehicle.save()
        return redirect('vehicles')

class VehicleUpdateView(UserPassesTestMixin,UpdateView):
    model = Vehicle
    form_class=VehicleEditForm
    context_object_name= 'vehicle'
    template_name = 'swmadmin/edit_vehicle.html'
    success_url = '/vehicles/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Vehicle = form.save(commit=False)
        Vehicle.updated_at = timezone.now()
        Vehicle.updated_by = self.request.user
        Vehicle.save()
        return redirect('vehicles')

class VehicleListView(UserPassesTestMixin,ListView):
    model = Vehicle
    context_object_name='vehicles'
    template_name = 'swmadmin/vehicles.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(VehicleListView,self).get_queryset().filter(is_active=True).order_by('plate_number')
        return qs

######################################Contractor#######################################################

class ContractorCreateView(UserPassesTestMixin,CreateView):
    model=Contractor
    form_class=NewContractorForm
    template_name = 'swmadmin/add_contractor.html'
    success_url= '/contractors'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Contractor = form.save(commit=False)
        Contractor.created_by = self.request.user
        Contractor.created_at = timezone.now()
        Contractor.save()
        return redirect('contractors')

class ContractorUpdateView(UserPassesTestMixin,UpdateView):
    model = Contractor
    form_class = ContractorEditForm
    context_object_name= 'Contractor'
    template_name = 'swmadmin/edit_contractor.html'
    success_url = '/contractors/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Contractor = form.save(commit=False)
        Contractor.updated_at = timezone.now()
        Contractor.updated_by = self.request.user
        Contractor.save()
        return redirect('contractors')

class ContractorListView(UserPassesTestMixin,ListView):
    model = Contractor
    context_object_name='contractors'
    template_name = 'swmadmin/contractors.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(ContractorListView,self).get_queryset().filter(is_active=True).order_by('name')
        return qs

####################################WCM##############################################################

class WCMCreateView(UserPassesTestMixin,CreateView):
    model=Ward_Contractor_Mapping
    form_class=NewWCMForm
    template_name = 'swmadmin/add_wcm.html'
    success_url= '/wcms'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        WCM = form.save(commit=False)
        WCM.created_by = self.request.user
        WCM.created_at = timezone.now()
        WCM.save()
        return redirect('wcms')

class WCMUpdateView(UserPassesTestMixin,UpdateView):
    model = Ward_Contractor_Mapping
    form_class = WCMEditForm
    context_object_name= 'wcm'
    template_name = 'swmadmin/edit_wcm.html'
    success_url = '/wcms/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        WCM = form.save(commit=False)
        WCM.updated_at = timezone.now()
        WCM.updated_by = self.request.user
        WCM.save()
        return redirect('wcms')

class WCMListView(UserPassesTestMixin,ListView):
    model = Ward_Contractor_Mapping
    context_object_name='wcms'
    template_name = 'swmadmin/wcms.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(WCMListView,self).get_queryset().filter(is_active=True)
        return qs

########################################VGM#########################################################

class VGMCreateView(UserPassesTestMixin,CreateView):
    model=Vehicle_Garage_Mapping
    form_class=NewVGMForm
    template_name = 'swmadmin/add_vgm.html'
    success_url= '/vgms'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        VGM = form.save(commit=False)
        VGM.created_by = self.request.user
        VGM.created_at = timezone.now()
        VGM.save()
        return redirect('vgms')

class VGMUpdateView(UserPassesTestMixin,UpdateView):
    model = Vehicle_Garage_Mapping
    form_class = VGMEditForm
    context_object_name= 'vgm'
    template_name = 'swmadmin/edit_vgm.html'
    success_url = '/vgms/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        VGM = form.save(commit=False)
        VGM.updated_at = timezone.now()
        VGM.updated_by = self.request.user
        VGM.save()
        return redirect('vgms')

class VGMListView(UserPassesTestMixin,ListView):
    model = Vehicle_Garage_Mapping
    context_object_name='vgms'
    template_name = 'swmadmin/vgms.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(VGMListView,self).get_queryset().filter(is_active=True)
        return qs

#######################################Installation##############################################

class InstallationCreateView(UserPassesTestMixin,CreateView):
    model=Installation
    form_class=NewInstallationForm
    template_name = 'swmadmin/add_installation.html'
    success_url= '/installations'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Installation = form.save(commit=False)
        Installation.created_by = self.request.user
        Installation.created_at = timezone.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        Installation.save()
        return redirect('installations')

class InstallationUpdateView(UserPassesTestMixin,UpdateView):
    model = Installation
    form_class=InstallationEditForm
    context_object_name= 'installation'
    template_name = 'swmadmin/edit_installation.html'
    success_url = '/installations/'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def form_valid(self,form):
        Installation = form.save(commit=False)
        Installation.updated_at = timezone.now()
        Installation.updated_by = self.request.user
        Installation.save()
        return redirect('installations')

class InstallationListView(UserPassesTestMixin,ListView):
    model = Installation
    context_object_name='installations'
    template_name = 'swmadmin/installations.html'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser

    def get_queryset(self):
        qs = super(InstallationListView,self).get_queryset().filter(is_active=True).order_by()
        return qs


####################################delete###########################################################
@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_vehicle(request):
    vehicle = Vehicle.objects.get(pk=request.POST['id'])
    vehicle.is_active = 'f'
    vehicle.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_bin(request):
    bn = Bin.objects.get(pk=request.POST['id'])
    bn.is_active = 'f'
    bn.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_stop_station(request):
    stop_station = Stop_station.objects.get(pk=request.POST['id'])
    stop_station.is_active = 'f'
    stop_station.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_route(request):
    route = Route.objects.get(pk=request.POST['id'])

    for bn in route.bins.all():
        bn.code = None
        bn.save()

    route.bins.clear()
    route.route_fence = None
    route.is_active = 'f'
    route.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_route_schedule(request):
    route_schedule = Route_schedule.objects.get(pk=request.POST['id'])
    route_schedule.is_active = 'f'
    route_schedule.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_contractor(request):
    contractor = Contractor.objects.get(pk=request.POST['id'])
    contractor.is_active = 'f'
    contractor.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_wcm(request):
    wcm = Ward_Contractor_Mapping.objects.get(pk=request.POST['id'])
    wcm.delete()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_vgm(request):
    vgm = Vehicle_Garage_Mapping.objects.get(pk=request.POST['id'])
    vgm.delete()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_installation(request):
    installation = Installation.objects.get(pk=request.POST['id'])
    installation.is_active = 'f'
    installation.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def delete_ewd(request):
    ewd = Ewd.objects.get(pk=request.POST['id'])
    ewd.is_active = 'f'
    ewd.save()
    response_data={}
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

####################################geofence area####################################################

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_bin_spot(request):
    bn = Bin.objects.get(pk=request.GET['id'])
    response_data={}
    response_data['bin_location'] = bn.bin_location.geojson
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_all_bin_spot(request):
    bns = Bin.objects.filter(is_active=True)
    locations = list();
    for bn in bns:
        location = list()
        location.append(bn.code)
        location.append(str(bn.longitude))
        location.append(str(bn.latitude))
        locations.append(location)

    response_data={}
    response_data['bin_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_all_bin_spot_from_ward(request):
    ward = Ward.objects.get(pk=request.GET['id'])
    bns = Bin.objects.filter(is_active=True).filter(bin_location__coveredby=ward.ward_fence)
    locations = list();
    for bn in bns:
        location = list()
        location.append(bn.code)
        location.append(str(bn.longitude))
        location.append(str(bn.latitude))
        locations.append(location)

    response_data={}
    response_data['bin_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
#@user_passes_test(lambda user: user.is_superuser)
def get_all_mlc_spot(request):
    mlcs = Stop_station.objects.filter(is_mlc=True)
    locations = list();
    for mlc in mlcs:
        location = list()
        location.append(mlc.name)
        mlc_centre = wkt.loads(str(mlc.stop_station_fence.wkt))
        location.append(str(mlc_centre.centroid.coords[0][0]))
        location.append(str(mlc_centre.centroid.coords[0][1]))
        locations.append(location)

    response_data={}
    response_data['mlc_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_all_tnsstn_spot(request):
    tnsstns = Stop_station.objects.filter(is_tnsstn=True)
    locations = list();
    for tnsstn in tnsstns:
        location = list()
        location.append(tnsstn.name)
        tnsstn_centre = wkt.loads(str(tnsstn.stop_station_fence.wkt))
        location.append(str(tnsstn_centre.centroid.coords[0][0]))
        location.append(str(tnsstn_centre.centroid.coords[0][1]))
        locations.append(location)

    response_data={}
    response_data['tnsstn_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_all_dmpgnd_spot(request):
    dmpgnds = Stop_station.objects.filter(is_dmpgnd=True)
    locations = list();
    for dmpgnd in dmpgnds:
        location = list()
        location.append(dmpgnd.name)
        dmpgnd_centre = wkt.loads(str(dmpgnd.stop_station_fence.wkt))
        location.append(str(dmpgnd_centre.centroid.coords[0][0]))
        location.append(str(dmpgnd_centre.centroid.coords[0][1]))
        locations.append(location)

    response_data={}
    response_data['dmpgnd_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_all_garage_spot(request):
    garages = Stop_station.objects.filter(is_garage=True)
    locations = list();
    for garage in garages:
        location = list()
        location.append(garage.name)
        garage_centre = wkt.loads(str(garage.stop_station_fence.wkt))
        location.append(str(garage_centre.centroid.coords[0][0]))
        location.append(str(garage_centre.centroid.coords[0][1]))
        locations.append(location)

    response_data={}
    response_data['garage_location'] = locations
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_stop_station_area(request):
    stop_station = Stop_station.objects.get(pk=request.GET['id'])
    response_data={}
    response_data['stop_station_fence'] = stop_station.stop_station_fence.geojson
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@user_passes_test(lambda user: user.is_superuser)
def get_ewd_area(request):
    ewd = Ewd.objects.get(pk=request.GET['id'])
    response_data={}
    response_data['ewd_fence'] = ewd.ewd_fence.geojson
    response_data['status'] = 'success'
    return HttpResponse(json.dumps(response_data),content_type="application/json")   
